from googleapiclient.discovery import build
from google.oauth2 import service_account
from openpyxl import Workbook
from openpyxl import load_workbook
import xlsxwriter
# from openpyxl import load_workbook
# import pandas as pd
import gspread
import win32com.client as win32
from win32com.client import Dispatch
# from gspread_dataframe import get_as_dataframe, set_with_dataframe
# from oauth2client.service_account import ServiceAccountCredentials

SERVICE_ACCOUNT_FILE = 'myproject4086.json'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']

cred = None
cred = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)

# The ID and range of a sample spreadsheet.
SAMPLE_SPREADSHEET_ID = '1rqsv6caQv-kveEz_FVaG8l2Dr-waJ9PRaJ-U0J-pN3w'

service = build('sheets', 'v4', credentials=cred)

# Call the Sheets API
sheet = service.spreadsheets()
print(sheet)
result = sheet.values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID, range="sheet1!A1:H16").execute()
values = result.get('values',[])
print(values)

with xlsxwriter.Workbook('new_sample4086.xlsx') as workbook:
    worksheet = workbook.add_worksheet()

    for row_num, data in enumerate(values):
        worksheet.write_row(row_num, 0, data)       
    
    

# import pandas as pd
# import gspread
# import win32com.client as win32
# from win32com.client import Dispatch
# from gspread_dataframe import get_as_dataframe, set_with_dataframe
# from oauth2client.service_account import ServiceAccountCredentials
# wb = load_workbook('updatedwarehouse.xlsx')
# wb.sheetnames
# ['data4086', 'sheet1']
# ws = wb.active
# # df = ws['B3:B11'].value
# test = ws.cell(row = 2, column= 4).value
# print(test)
# byteData = sheet(fileID = SAMPLE_SPREADSHEET_ID, mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet').execute()
# with open('updatwarehouse.xlsx', 'wb') as f:
# f.write(byteData)
# f.close
# from openpyxl import load_workbook
# ws.cell(row=3, column=3).value
# <Worksheet "Sheet1">
# xlApp = win32.Dispatch('Excel.Application')
# wb =Workbook("updatedwarehouse.xlsx")
# Workbook.active
# print(wb)
# ws = wb.create_sheet(title= 'data4', index= 1)
# print(ws)
# ws.append(result)
# rowNumber = 1
# colCount= len(rows[0])
# ws.update_a_cell('A1', 'Gspread !')
# for row in values:
#     wb.range(wb.cells(rowNumber, 1), wb.cells(rowNumber, colCount)).value = row 
#     rowNumber += 1